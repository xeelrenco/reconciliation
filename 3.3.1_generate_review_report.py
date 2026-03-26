"""
renco_export_v2.py
==================
Genera l'Excel MDR riconciliazione con esattamente le colonne
del file originale + motivazioni GPT e Claude.

Colonne output:
  # | Titolo MDR (storico) | N. Documento MDR | Progetto MDR |
  Esito | Titolo RACI Abbinato | Confidence |
  Motivazione della Scelta | Motivazione GPT | Motivazione Claude

Dipendenze:
    pip install duckdb openpyxl

Configurazione:
    Preferito: inserisci MOTHERDUCK_TOKEN (e opzionale MOTHERDUCK_DB) in config.txt
    nella stessa cartella di questo script.
    Fallback: variabili d'ambiente MOTHERDUCK_TOKEN / MOTHERDUCK_DB.

Uso:
    python 3.3.1_generate_review_report.py
"""

import os
import re
from pathlib import Path
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────
PROMPT_VERSION   = "v1.2"
OUTPUT_FILE      = "renco_riconciliazione_report.xlsx"

CONFIG_PATH = Path(__file__).resolve().parent / "config.txt"


def load_config(path: Path = CONFIG_PATH) -> dict:
    cfg = {}
    if not path.exists():
        return cfg
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, _, v = line.partition("=")
                cfg[k.strip()] = v.strip()
    return cfg


CFG = load_config()
MOTHERDUCK_TOKEN = CFG.get("MOTHERDUCK_TOKEN") or os.environ.get("MOTHERDUCK_TOKEN", "")
MOTHERDUCK_DB = CFG.get("MOTHERDUCK_DB") or os.environ.get("MOTHERDUCK_DB", "my_db")
# Connessione MotherDuck: usa token se disponibile (config.txt o env), altrimenti prova senza token.
CONN_STR = f"md:{MOTHERDUCK_DB}?motherduck_token={MOTHERDUCK_TOKEN}" if MOTHERDUCK_TOKEN else f"md:{MOTHERDUCK_DB}"


def _norm(s):
    """Normalizza titolo: lowercase + collassa tutti i whitespace a singolo spazio."""
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


# ──────────────────────────────────────────────────────────────
# QUERIES
# ──────────────────────────────────────────────────────────────
SQL_RESULTS = f"""
SELECT
    t.Document_title                                              AS titolo_mdr,
    r.FinalDecisionType                                           AS esito,
    COALESCE(r.FinalRaciTitle, '—')                               AS raci_abbinato,
    ROUND(r.FinalConfidence, 3)                                   AS confidence,
    r.FinalReason                                                 AS motivazione_finale,
    MAX(CASE WHEN d.AgentName = 'gpt5mini' THEN d.ReasoningSummary END) AS motivazione_gpt,
    MAX(CASE WHEN d.AgentName = 'claude'   THEN d.ReasoningSummary END) AS motivazione_claude
FROM my_db.mdr_reconciliation.MdrReconciliationResults r
JOIN my_db.mdr_reconciliation.MdrReconciliationTasks t
    ON t.TaskId = r.TaskId
LEFT JOIN my_db.mdr_reconciliation.MdrReconciliationAgentDecisions d
    ON d.TaskId = r.TaskId
   AND d.AgentName IN ('gpt5mini', 'claude')
WHERE r.PromptVersion = '{PROMPT_VERSION}'
GROUP BY
    t.Document_title, r.FinalDecisionType, r.FinalRaciTitle,
    r.FinalConfidence, r.FinalReason
ORDER BY r.FinalDecisionType, t.Document_title
"""

SQL_RAW = """
SELECT
    Document_title,
    Document_number,
    Mdr_code_name_ref
FROM my_db.historical_mdr_normalization.MdrPreviousRecordsRaw
ORDER BY Document_number
"""


def build_raw_lookup(raw_rows):
    """
    Costruisce dict {titolo_normalizzato -> (Document_number, Mdr_code_name_ref)}.
    Normalizzazione in Python con re.sub — gestisce spazi multipli, newline, tab.
    Mantiene solo il primo match per titolo (ORDER BY Document_number garantisce
    che sia il numero più basso/alfabetico).
    """
    lookup = {}
    for title, doc_num, progetto in raw_rows:
        key = _norm(title)
        if key and key not in lookup:
            lookup[key] = (doc_num, progetto)
    return lookup


# ──────────────────────────────────────────────────────────────
# STILI
# ──────────────────────────────────────────────────────────────
NAVY   = "0D1B2A"
WHITE  = "FFFFFF"
BORDER = "CBD5E1"

ESITO_BG = {"MATCH": "D1FAE5", "NO_MATCH": "FEE2E2", "MANUAL_REVIEW": "FEF3C7"}
ESITO_FG = {"MATCH": "065F46", "NO_MATCH": "7F1D1D", "MANUAL_REVIEW": "92400E"}

thin      = Side(style="thin",   color=BORDER)
thick_top = Side(style="medium", color="334155")


def _border(top=False):
    t = thick_top if top else thin
    return Border(left=thin, right=thin, top=t, bottom=thin)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)


def _font(bold=False, color="1A1A2E", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)


def _align(h="left", wrap=True, v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ──────────────────────────────────────────────────────────────
# BUILD EXCEL
# ──────────────────────────────────────────────────────────────
HEADERS = [
    "#",
    "Titolo MDR (storico)",
    "N. Documento MDR",
    "Progetto MDR",
    "Esito",
    "Titolo RACI Abbinato",
    "Confidence",
    "Motivazione della Scelta Finale(giudice Gemini)",
    "Motivazione Agente1 GPT",
    "Motivazione Agente2 Claude",
]

COL_WIDTHS = [5, 68, 28, 32, 14, 58, 12, 70, 70, 70]

SHEET_CONFIGS = [
    ("Riconciliazione Completa", None),
    ("MATCH",         "MATCH"),
    ("NO_MATCH",      "NO_MATCH"),
    ("MANUAL REVIEW", "MANUAL_REVIEW"),
]


def build_excel(rows):
    wb = Workbook()
    first = True
    for sheet_name, esito_filter in SHEET_CONFIGS:
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        filtered = [r for r in rows if esito_filter is None or r["esito"] == esito_filter]
        _build_sheet(ws, filtered, sheet_name)

    wb.save(OUTPUT_FILE)
    print(f"\n✓  Salvato: {OUTPUT_FILE}")
    for _, ef in SHEET_CONFIGS:
        n = len([r for r in rows if ef is None or r["esito"] == ef])
        print(f"   {ef or 'TOTALE'}: {n} righe")


def _build_sheet(ws, rows, title):
    total = len(rows)

    # Riga 1 — titolo foglio
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    ws["A1"] = f"Renco MDR — {title}  |  {total} documenti  |  Prompt {PROMPT_VERSION}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws["A1"].fill      = _fill(NAVY)
    ws["A1"].alignment = _align(h="left", v="center", wrap=False)
    ws.row_dimensions[1].height = 26

    # Riga 2 — header colonne
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill      = _fill("1A2E42")
        c.alignment = _align(h="center", v="center", wrap=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # Dati
    prev_esito = None
    for idx, row in enumerate(rows, 1):
        r     = idx + 2
        esito = row["esito"]
        conf  = row["confidence"]
        bg    = ESITO_BG.get(esito, "F8FAFC")
        fg    = ESITO_FG.get(esito, "1A1A2E")
        top   = prev_esito is not None and prev_esito != esito

        vals = [
            idx,
            row["titolo_mdr"],
            row["numero_documento"] or "—",
            row["progetto"]         or "—",
            esito,
            row["raci_abbinato"],
            conf,
            row["motivazione_finale"] or "—",
            row["motivazione_gpt"]    or "—",
            row["motivazione_claude"] or "—",
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.border    = _border(top)
            c.fill      = _fill(bg)
            c.alignment = _align(
                h="center" if ci in (1, 5, 7) else "left",
                wrap=True, v="top",
            )
            c.font = (
                Font(name="Arial", size=9, bold=True, color=fg)
                if ci == 5
                else _font(size=9)
            )
            if ci == 7 and conf is not None:
                c.number_format = "0.000"

        ws.row_dimensions[r].height = 60
        prev_esito = esito

    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{total + 2}"
    ws.freeze_panes    = "A3"


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────
def main():
    print("Connessione a MotherDuck...")
    con = duckdb.connect(CONN_STR)

    print("Query risultati riconciliazione...")
    raw_results = con.execute(SQL_RESULTS).fetchall()
    col_names = [
        "titolo_mdr", "esito", "raci_abbinato", "confidence",
        "motivazione_finale", "motivazione_gpt", "motivazione_claude",
    ]
    rows = [dict(zip(col_names, r)) for r in raw_results]
    print(f"  {len(rows)} task recuperati.")

    print("Query tabella raw per numeri documento e progetti...")
    raw_raw = con.execute(SQL_RAW).fetchall()
    lookup  = build_raw_lookup(raw_raw)
    print(f"  {len(lookup)} titoli unici nella raw.")
    con.close()

    # Arricchisci ogni riga con numero documento e progetto (join in Python)
    not_found = 0
    for row in rows:
        key = _norm(row["titolo_mdr"])
        doc_num, progetto = lookup.get(key, (None, None))
        row["numero_documento"] = doc_num
        row["progetto"]         = progetto
        if doc_num is None:
            not_found += 1

    if not_found:
        print(f"  ⚠  {not_found} titoli senza corrispondenza nella raw.")
    else:
        print("  ✓  Tutti i titoli hanno trovato numero documento e progetto.")

    print("Generazione Excel...")
    build_excel(rows)


if __name__ == "__main__":
    main()