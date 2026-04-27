"""
3.4.1_generate_review_with_top3_and_recovery.py
================================================
Genera l'Excel MDR riconciliazione con motivazioni, scelta agente, top 3 candidati
per GPT e Claude, includendo anche i risultati del recovery agent (script 3.4).

Colonne output (aggiunte recovery):
  - Recovery 3.4 — Nuovo Esito
  - Recovery 3.4 — RACI Title scelto
  - Recovery 3.4 — Motivazione

Dipendenze:
    pip install duckdb openpyxl

Configurazione:
    Preferito: inserisci MOTHERDUCK_TOKEN (e opzionale MOTHERDUCK_DB) in config.txt
    nella stessa cartella di questo script.
    Fallback: variabili d'ambiente MOTHERDUCK_TOKEN / MOTHERDUCK_DB.

Uso:
    python 3.4.1_generate_review_with_top3_and_recovery.py
"""

import os
import re
from pathlib import Path
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────
PROMPT_VERSION = "v1.2"
OUTPUT_FILE = "renco_riconciliazione_report_with_top3_and_recovery.xlsx"

CONFIG_PATH = Path(__file__).resolve().parent / "config.txt"


def load_config(path: Path = CONFIG_PATH) -> dict:
    cfg = {}
    if not path.exists():
        return cfg
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, _, v = line.partition("=")
                cfg[k.strip()] = v.strip()
    return cfg


CFG = load_config()
MOTHERDUCK_TOKEN = CFG.get("MOTHERDUCK_TOKEN") or os.environ.get("MOTHERDUCK_TOKEN", "")
MOTHERDUCK_DB = CFG.get("MOTHERDUCK_DB") or os.environ.get("MOTHERDUCK_DB", "my_db")
CONN_STR = f"md:{MOTHERDUCK_DB}?motherduck_token={MOTHERDUCK_TOKEN}" if MOTHERDUCK_TOKEN else f"md:{MOTHERDUCK_DB}"


def _norm(s):
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _safe_excel_value(value):
    """
    Rimuove caratteri non ammessi da openpyxl/XML e limita la lunghezza stringa.
    """
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_XLSX_CHARS_RE.sub("", value)
    return cleaned[:32767]


# ──────────────────────────────────────────────────────────────
# QUERIES
# ──────────────────────────────────────────────────────────────
SQL_RESULTS = f"""
SELECT
    t.TaskId,
    t.Document_title                                              AS titolo_mdr,
    r.FinalDecisionType                                           AS esito,
    COALESCE(r.FinalRaciTitle, '—')                              AS raci_abbinato,
    ROUND(r.FinalConfidence, 3)                                   AS confidence,
    r.FinalReason                                                 AS motivazione_finale,
    MAX(CASE WHEN d.AgentName = 'gpt5mini' THEN d.ReasoningSummary    END) AS motivazione_gpt,
    MAX(CASE WHEN d.AgentName = 'claude'   THEN d.ReasoningSummary    END) AS motivazione_claude,
    MAX(CASE WHEN d.AgentName = 'gpt5mini' THEN d.DecisionType        END) AS gpt_decision,
    MAX(CASE WHEN d.AgentName = 'gpt5mini' THEN d.SelectedRaciTitle   END) AS gpt_selected_raci,
    MAX(CASE WHEN d.AgentName = 'gpt5mini' THEN ROUND(d.Confidence,3) END) AS gpt_conf,
    MAX(CASE WHEN d.AgentName = 'claude'   THEN d.DecisionType        END) AS claude_decision,
    MAX(CASE WHEN d.AgentName = 'claude'   THEN d.SelectedRaciTitle   END) AS claude_selected_raci,
    MAX(CASE WHEN d.AgentName = 'claude'   THEN ROUND(d.Confidence,3) END) AS claude_conf,
    rr.RecoveryDecisionType                                        AS recovery_decision,
    rr.RecoveryRaciTitle                                           AS recovery_raci_title,
    rr.RecoveryReason                                              AS recovery_reason
FROM my_db.mdr_reconciliation.MdrReconciliationResults r
JOIN my_db.mdr_reconciliation.MdrReconciliationTasks t
    ON t.TaskId = r.TaskId
LEFT JOIN my_db.mdr_reconciliation.MdrReconciliationAgentDecisions d
    ON d.TaskId = r.TaskId
   AND d.AgentName IN ('gpt5mini', 'claude')
LEFT JOIN my_db.mdr_reconciliation.MdrReconciliationRecoveryResults rr
    ON rr.TaskId = r.TaskId
   AND rr.PromptVersion = r.PromptVersion
   AND rr.EmbeddingModel = r.EmbeddingModel
   AND rr.RecoveryStage = CASE
         WHEN r.FinalDecisionType = 'MANUAL_REVIEW' THEN 'manual_review_resolver'
         WHEN r.FinalDecisionType = 'NO_MATCH' THEN 'no_match_recovery'
         ELSE '__unsupported__'
       END
WHERE r.PromptVersion = '{PROMPT_VERSION}'
GROUP BY
    t.TaskId, t.Document_title, r.FinalDecisionType, r.FinalRaciTitle,
    r.FinalConfidence, r.FinalReason, rr.RecoveryDecisionType, rr.RecoveryRaciTitle, rr.RecoveryReason
ORDER BY r.FinalDecisionType, t.Document_title
"""

SQL_CANDIDATES = f"""
SELECT
    TaskId,
    AgentName,
    CandidateRankWithinAgent                    AS rank,
    COALESCE(RaciTitle, '—')                    AS raci_title,
    ROUND(CandidateConfidence, 3)               AS conf,
    COALESCE(WhyPlausible, '')                  AS why
FROM my_db.mdr_reconciliation.MdrReconciliationAgentTopCandidates
WHERE PromptVersion = '{PROMPT_VERSION}'
  AND CandidateRankWithinAgent <= 3
ORDER BY TaskId, AgentName, CandidateRankWithinAgent
"""

SQL_RAW = """
SELECT Document_title, Document_number, Mdr_code_name_ref
FROM my_db.historical_mdr_normalization.MdrPreviousRecordsRaw
ORDER BY Document_number
"""


def build_raw_lookup(raw_rows):
    lookup = {}
    for title, doc_num, progetto in raw_rows:
        key = _norm(title)
        if key and key not in lookup:
            lookup[key] = (doc_num, progetto)
    return lookup


def build_candidate_lookup(cand_rows):
    lookup = {}
    for task_id, agent, rank, raci_title, conf, why in cand_rows:
        agent_key = "gpt" if agent == "gpt5mini" else "claude"
        if task_id not in lookup:
            lookup[task_id] = {"gpt": [], "claude": []}
        lookup[task_id][agent_key].append((raci_title, conf, why))
    return lookup


def fmt_candidate(triple):
    if not triple:
        return "—"
    raci_title, conf, why = triple
    why_str = (why or "").strip()
    title = (raci_title or "—").strip()
    if not why_str:
        return f"TITOLO:\n{title}\n\nNOTA:\n—"
    return f"TITOLO:\n{title}\n\nNOTA:\n{why_str}"


def fmt_decision(decision, raci, conf):
    if not decision:
        return "—"
    if decision == "NO_MATCH":
        return "DECISIONE: NO MATCH\n\nNessun titolo selezionato"
    raci_str = raci or "—"
    return f"DECISIONE: MATCH\n\nTITOLO SCELTO:\n{raci_str}"


# ──────────────────────────────────────────────────────────────
# STILI
# ──────────────────────────────────────────────────────────────
NAVY = "0D1B2A"
WHITE = "FFFFFF"
BORDER = "CBD5E1"

ESITO_BG = {"MATCH": "D1FAE5", "NO_MATCH": "FEE2E2", "MANUAL_REVIEW": "FEF3C7"}
ESITO_FG = {"MATCH": "065F46", "NO_MATCH": "7F1D1D", "MANUAL_REVIEW": "92400E"}

GPT_BG = "EFF6FF"
CLAUDE_BG = "F5F3FF"
RECOVERY_BG = "ECFDF5"

thin = Side(style="thin", color=BORDER)
thick_top = Side(style="medium", color="334155")


def _border(top=False):
    t = thick_top if top else thin
    return Border(left=thin, right=thin, top=t, bottom=thin)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)


def _font(bold=False, color="1A1A2E", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)


def _align(h="left", wrap=True, v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ──────────────────────────────────────────────────────────────
# STRUTTURA COLONNE
# ──────────────────────────────────────────────────────────────
HEADERS = [
    "#",
    "Titolo MDR (storico)",
    "N. Documento MDR",
    "Progetto MDR",
    "Esito",
    "Titolo RACI Abbinato",
    "Motivazione della Scelta Finale(giudice Gemini)",
    "GPT — Scelta",
    "Motivazione GPT",
    "GPT — CandidatoTop 1",
    "GPT — Candidato Top 2",
    "GPT — Candidato Top 3",
    "Claude — Scelta",
    "Motivazione Claude",
    "Claude — Candidato Top 1",
    "Claude — Candidato Top 2",
    "Claude — Candidato Top 3",
    "Recovery 3.4 — Nuovo Esito",
    "Recovery 3.4 — RACI Title scelto",
    "Recovery 3.4 — Motivazione",
]

COL_WIDTHS = [
    5, 68, 28, 32, 14, 58, 70,
    45, 70, 45, 45, 45,
    45, 70, 45, 45, 45,
    24, 58, 70,
]

GPT_COLS = {8, 9, 10, 11, 12}
CLAUDE_COLS = {13, 14, 15, 16, 17}
RECOVERY_COLS = {18, 19, 20}

SHEET_CONFIGS = [
    ("Riconciliazione Completa", None),
    ("MATCH", "MATCH"),
    ("NO_MATCH", "NO_MATCH"),
    ("MANUAL REVIEW", "MANUAL_REVIEW"),
]


def build_excel(rows):
    wb = Workbook()
    first = True
    for sheet_name, esito_filter in SHEET_CONFIGS:
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        filtered = [r for r in rows if esito_filter is None or r["esito"] == esito_filter]
        _build_sheet(ws, filtered, sheet_name)

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Salvato: {OUTPUT_FILE}")
    for _, ef in SHEET_CONFIGS:
        n = len([r for r in rows if ef is None or r["esito"] == ef])
        print(f"   {ef or 'TOTALE'}: {n} righe")


def _build_sheet(ws, rows, title):
    total = len(rows)

    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    ws["A1"] = f"Renco MDR — {title}  |  {total} documenti  |  Prompt {PROMPT_VERSION}"
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = _align(h="left", v="center", wrap=False)
    ws.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=2, column=ci, value=h)
        if ci in GPT_COLS:
            c.fill = _fill("1E40AF")
        elif ci in CLAUDE_COLS:
            c.fill = _fill("5B21B6")
        elif ci in RECOVERY_COLS:
            c.fill = _fill("065F46")
        else:
            c.fill = _fill("1A2E42")
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.alignment = _align(h="center", v="center", wrap=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    prev_esito = None
    for idx, row in enumerate(rows, 1):
        r = idx + 2
        esito = row["esito"]
        bg = ESITO_BG.get(esito, "F8FAFC")
        fg = ESITO_FG.get(esito, "1A1A2E")
        top = prev_esito is not None and prev_esito != esito

        gpt_cands = row.get("gpt_candidates", [None, None, None])
        claude_cands = row.get("claude_candidates", [None, None, None])

        vals = [
            idx,
            row["titolo_mdr"],
            row["numero_documento"] or "—",
            row["progetto"] or "—",
            esito,
            row["raci_abbinato"],
            row["motivazione_finale"] or "—",
            fmt_decision(row["gpt_decision"], row["gpt_selected_raci"], row["gpt_conf"]),
            row["motivazione_gpt"] or "—",
            fmt_candidate(gpt_cands[0] if len(gpt_cands) > 0 else None),
            fmt_candidate(gpt_cands[1] if len(gpt_cands) > 1 else None),
            fmt_candidate(gpt_cands[2] if len(gpt_cands) > 2 else None),
            fmt_decision(row["claude_decision"], row["claude_selected_raci"], row["claude_conf"]),
            row["motivazione_claude"] or "—",
            fmt_candidate(claude_cands[0] if len(claude_cands) > 0 else None),
            fmt_candidate(claude_cands[1] if len(claude_cands) > 1 else None),
            fmt_candidate(claude_cands[2] if len(claude_cands) > 2 else None),
            row.get("recovery_decision") or "—",
            row.get("recovery_raci_title") or "—",
            row.get("recovery_reason") or "—",
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=_safe_excel_value(val))
            c.border = _border(top)

            if ci in GPT_COLS:
                c.fill = _fill(GPT_BG)
                c.font = _font(size=9, color="1E3A8A")
            elif ci in CLAUDE_COLS:
                c.fill = _fill(CLAUDE_BG)
                c.font = _font(size=9, color="4C1D95")
            elif ci in RECOVERY_COLS:
                c.fill = _fill(RECOVERY_BG)
                c.font = _font(size=9, color="065F46")
            elif ci == 5:
                c.fill = _fill(bg)
                c.font = Font(name="Arial", size=9, bold=True, color=fg)
            else:
                c.fill = _fill(bg)
                c.font = _font(size=9)

            c.alignment = _align(
                h="center" if ci in (1, 5, 18) else "left",
                wrap=True, v="top",
            )

        ws.row_dimensions[r].height = 95
        prev_esito = esito

    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{total + 2}"
    ws.freeze_panes = "B3"


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────
def main():
    print("Connessione a MotherDuck...")
    con = duckdb.connect(CONN_STR)

    print("Query risultati riconciliazione + recovery...")
    raw_results = con.execute(SQL_RESULTS).fetchall()
    col_names = [
        "task_id", "titolo_mdr", "esito", "raci_abbinato", "confidence",
        "motivazione_finale", "motivazione_gpt", "motivazione_claude",
        "gpt_decision", "gpt_selected_raci", "gpt_conf",
        "claude_decision", "claude_selected_raci", "claude_conf",
        "recovery_decision", "recovery_raci_title", "recovery_reason",
    ]
    rows = [dict(zip(col_names, r)) for r in raw_results]
    print(f"  {len(rows)} task recuperati.")

    print("Query top candidati per agente...")
    cand_raw = con.execute(SQL_CANDIDATES).fetchall()
    cand_look = build_candidate_lookup(cand_raw)
    print(f"  {len(cand_look)} task con candidati.")

    print("Query tabella raw per numeri documento e progetti...")
    raw_raw = con.execute(SQL_RAW).fetchall()
    raw_look = build_raw_lookup(raw_raw)
    print(f"  {len(raw_look)} titoli unici nella raw.")
    con.close()

    not_found = 0
    for row in rows:
        key = _norm(row["titolo_mdr"])
        doc_num, progetto = raw_look.get(key, (None, None))
        row["numero_documento"] = doc_num
        row["progetto"] = progetto
        if doc_num is None:
            not_found += 1

        task_cands = cand_look.get(row["task_id"], {"gpt": [], "claude": []})
        row["gpt_candidates"] = task_cands["gpt"]
        row["claude_candidates"] = task_cands["claude"]

    if not_found:
        print(f"  [WARN] {not_found} titoli senza corrispondenza nella raw.")
    else:
        print("  [OK] Tutti i titoli hanno trovato numero documento e progetto.")

    print("Generazione Excel...")
    build_excel(rows)


if __name__ == "__main__":
    main()
